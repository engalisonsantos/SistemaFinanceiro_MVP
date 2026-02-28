from datetime import date
from decimal import Decimal
from io import BytesIO

from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.views.decorators.http import require_http_methods

from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .models import Group, Subgroup, Transaction
from .decorators import master_required


def home_redirect(request):
    return redirect("transactions")


@login_required
@require_http_methods(["GET", "POST"])
def transactions_view(request):
    if request.method == "POST":
        tx = Transaction(created_by=request.user)
        tx.date = request.POST["date"]
        tx.group_id = request.POST["group"]
        tx.subgroup_id = request.POST["subgroup"]
        tx.name = request.POST["name"].strip()
        tx.description = request.POST.get("description", "").strip()
        tx.value = request.POST["value"]

        # trava: subgrupo precisa pertencer ao grupo selecionado
        sg = get_object_or_404(Subgroup, pk=tx.subgroup_id)
        if str(sg.group_id) != str(tx.group_id):
            d1 = request.GET.get("d1") or date.today().replace(day=1).isoformat()
            d2 = request.GET.get("d2") or date.today().isoformat()
            f_group = request.GET.get("group") or ""
            f_subgroup = request.GET.get("subgroup") or ""
            q = request.GET.get("q") or ""

            qs = Transaction.objects.select_related("group", "subgroup") \
                .filter(date__range=[d1, d2]).order_by("-date", "-id")

            if f_group:
                qs = qs.filter(group_id=f_group)
            if f_subgroup:
                qs = qs.filter(subgroup_id=f_subgroup)
            if q:
                qs = qs.filter(name__icontains=q) | qs.filter(description__icontains=q)

            context = {
                "error": "Subgrupo inválido para o Grupo selecionado. Selecione novamente.",
                "groups": Group.objects.all().order_by("name"),
                "subgroups": Subgroup.objects.select_related("group").all().order_by("group__name", "name"),
                "transactions": qs[:500],
                "d1": d1, "d2": d2,
                "f_group": f_group, "f_subgroup": f_subgroup, "q": q,
            }
            return render(request, "finance/transactions.html", context)

        tx.save()
        return redirect("transactions")

    d1 = request.GET.get("d1") or date.today().replace(day=1).isoformat()
    d2 = request.GET.get("d2") or date.today().isoformat()
    f_group = request.GET.get("group") or ""
    f_subgroup = request.GET.get("subgroup") or ""
    q = request.GET.get("q") or ""

    qs = Transaction.objects.select_related("group", "subgroup").filter(date__range=[d1, d2]).order_by("-date", "-id")
    if f_group:
        qs = qs.filter(group_id=f_group)
    if f_subgroup:
        qs = qs.filter(subgroup_id=f_subgroup)
    if q:
        qs = qs.filter(name__icontains=q) | qs.filter(description__icontains=q)

    context = {
        "groups": Group.objects.all().order_by("name"),
        "subgroups": Subgroup.objects.select_related("group").all().order_by("group__name", "name"),
        "transactions": qs[:500],
        "d1": d1, "d2": d2,
        "f_group": f_group, "f_subgroup": f_subgroup, "q": q,
    }
    return render(request, "finance/transactions.html", context)


@login_required
@require_http_methods(["POST"])
def transaction_delete(request, pk: int):
    tx = get_object_or_404(Transaction, pk=pk)
    tx.delete()
    return redirect("transactions")


@login_required
def subgroups_by_group(request, group_id: int):
    data = list(Subgroup.objects.filter(group_id=group_id).order_by("name").values("id", "name"))
    return JsonResponse({"items": data})


@login_required
@master_required
def report_period(request):
    d1 = request.GET.get("d1") or date.today().replace(day=1).isoformat()
    d2 = request.GET.get("d2") or date.today().isoformat()

    qs = Transaction.objects.select_related("group", "subgroup").filter(date__range=[d1, d2]).order_by("date", "id")

    receipt_total = qs.filter(type=Transaction.Type.RECEIPT).aggregate(s=Sum("value"))["s"] or Decimal("0.00")
    payment_total = qs.filter(type=Transaction.Type.PAYMENT).aggregate(s=Sum("value"))["s"] or Decimal("0.00")
    balance = receipt_total - payment_total

    fmt = request.GET.get("fmt") or "html"  # html | xlsx | pdf

    if fmt == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório"

        ws.append(["Período", f"{d1} a {d2}"])
        ws.append(["Recebimentos", float(receipt_total)])
        ws.append(["Pagamentos", float(payment_total)])
        ws.append(["Saldo", float(balance)])
        ws.append([])
        ws.append(["Data", "Tipo", "Grupo", "Subgrupo", "Nome", "Descrição", "Valor"])

        for tx in qs:
            ws.append([
                tx.date.isoformat(),
                tx.get_type_display(),
                tx.group.name,
                tx.subgroup.name,
                tx.name,
                tx.description,
                float(tx.value),
            ])

        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 22

        resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = f'attachment; filename="relatorio_{d1}_a_{d2}.xlsx"'
        wb.save(resp)
        return resp

    if fmt == "pdf":
        buffer = BytesIO()
        c = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4

        y = height - 18*mm
        c.setFont("Helvetica-Bold", 14)
        c.drawString(15*mm, y, "Relatório por Período")
        y -= 8*mm

        c.setFont("Helvetica", 10)
        c.drawString(15*mm, y, f"Período: {d1} a {d2}")
        y -= 8*mm

        c.setFont("Helvetica-Bold", 11)
        c.drawString(15*mm, y, "Resumo")
        y -= 6*mm

        c.setFont("Helvetica", 10)
        c.drawString(15*mm, y, f"Recebimentos: R$ {receipt_total:.2f}")
        y -= 5*mm
        c.drawString(15*mm, y, f"Pagamentos:   R$ {payment_total:.2f}")
        y -= 5*mm
        c.drawString(15*mm, y, f"Saldo:        R$ {balance:.2f}")
        y -= 10*mm

        c.setFont("Helvetica-Bold", 9)
        headers = ["Data", "Tipo", "Grupo/Subgrupo", "Nome", "Valor (R$)"]
        col_x = [15*mm, 35*mm, 70*mm, 135*mm, 200*mm]

        for i, h in enumerate(headers):
            c.drawString(col_x[i], y, h)
        y -= 4*mm
        c.line(15*mm, y, 200*mm, y)
        y -= 6*mm

        c.setFont("Helvetica", 9)

        def new_page():
            nonlocal y
            c.showPage()
            y = height - 18*mm
            c.setFont("Helvetica-Bold", 12)
            c.drawString(15*mm, y, "Relatório por Período (continuação)")
            y -= 10*mm
            c.setFont("Helvetica-Bold", 9)
            for i, h in enumerate(headers):
                c.drawString(col_x[i], y, h)
            y -= 4*mm
            c.line(15*mm, y, 200*mm, y)
            y -= 6*mm
            c.setFont("Helvetica", 9)

        for tx in qs:
            if y < 20*mm:
                new_page()

            dt = tx.date.strftime("%d/%m/%Y")
            tipo = tx.get_type_display()
            grp = f"{tx.group.name}/{tx.subgroup.name}"
            nome = (tx.name or "")[:28]
            valor = f"{tx.value:.2f}"

            c.drawString(col_x[0], y, dt)
            c.drawString(col_x[1], y, tipo)
            c.drawString(col_x[2], y, grp[:35])
            c.drawString(col_x[3], y, nome)
            c.drawRightString(col_x[4], y, valor)
            y -= 5*mm

        c.save()
        buffer.seek(0)

        resp = HttpResponse(buffer.getvalue(), content_type="application/pdf")
        resp["Content-Disposition"] = f'attachment; filename="relatorio_{d1}_a_{d2}.pdf"'
        return resp

    return render(request, "finance/report_period.html", {
        "d1": d1, "d2": d2,
        "transactions": qs,
        "receipt_total": receipt_total,
        "payment_total": payment_total,
        "balance": balance,
    })


@login_required
@master_required
def report_dre(request):
    """
    DRE simples e automática:
    - Agrupa por NATUREZA (Receita/Despesa) e por NOME DO GRUPO
    - Não existe dre_section
    """
    d1 = request.GET.get("d1") or date.today().replace(day=1).isoformat()
    d2 = request.GET.get("d2") or date.today().isoformat()

    qs = Transaction.objects.select_related("group").filter(date__range=[d1, d2])

    receita_total = qs.filter(group__nature=Group.Nature.RECEITA).aggregate(s=Sum("value"))["s"] or Decimal("0.00")
    despesa_total = qs.filter(group__nature=Group.Nature.DESPESA).aggregate(s=Sum("value"))["s"] or Decimal("0.00")
    resultado = receita_total - despesa_total

    # Totais por grupo (separando por natureza)
    receitas_por_grupo = (
        qs.filter(group__nature=Group.Nature.RECEITA)
        .values("group__name")
        .annotate(total=Sum("value"))
        .order_by("-total", "group__name")
    )

    despesas_por_grupo = (
        qs.filter(group__nature=Group.Nature.DESPESA)
        .values("group__name")
        .annotate(total=Sum("value"))
        .order_by("-total", "group__name")
    )

    fmt = request.GET.get("fmt") or "html"  # html | xlsx | pdf

    if fmt == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "DRE"

        ws.append(["DRE (automática por Grupo)"])
        ws.append(["Período", f"{d1} a {d2}"])
        ws.append([])
        ws.append(["Receita total", float(receita_total)])
        ws.append(["Despesa total", float(despesa_total)])
        ws.append(["Resultado", float(resultado)])
        ws.append([])

        ws.append(["RECEITAS (por grupo)"])
        ws.append(["Grupo", "Total"])
        for r in receitas_por_grupo:
            ws.append([r["group__name"], float(r["total"] or 0)])

        ws.append([])
        ws.append(["DESPESAS (por grupo)"])
        ws.append(["Grupo", "Total"])
        for r in despesas_por_grupo:
            ws.append([r["group__name"], float(r["total"] or 0)])

        ws.column_dimensions["A"].width = 45
        ws.column_dimensions["B"].width = 18

        resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = f'attachment; filename="dre_{d1}_a_{d2}.xlsx"'
        wb.save(resp)
        return resp

    if fmt == "pdf":
        buffer = BytesIO()
        c = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4

        y = height - 18 * mm
        c.setFont("Helvetica-Bold", 14)
        c.drawString(15 * mm, y, "DRE (automática por Grupo)")
        y -= 8 * mm

        c.setFont("Helvetica", 10)
        c.drawString(15 * mm, y, f"Período: {d1} a {d2}")
        y -= 10 * mm

        c.setFont("Helvetica-Bold", 11)
        c.drawString(15 * mm, y, "Resumo")
        y -= 6 * mm
        c.setFont("Helvetica", 10)
        c.drawString(15 * mm, y, f"Receita total: R$ {receita_total:.2f}")
        y -= 5 * mm
        c.drawString(15 * mm, y, f"Despesa total: R$ {despesa_total:.2f}")
        y -= 5 * mm
        c.drawString(15 * mm, y, f"Resultado:    R$ {resultado:.2f}")
        y -= 10 * mm

        def ensure_space(min_y=25 * mm):
            nonlocal y
            if y < min_y:
                c.showPage()
                y = height - 18 * mm

        c.setFont("Helvetica-Bold", 11)
        c.drawString(15 * mm, y, "Receitas por Grupo")
        y -= 7 * mm
        c.setFont("Helvetica-Bold", 9)
        c.drawString(15 * mm, y, "Grupo")
        c.drawRightString(200 * mm, y, "Total (R$)")
        y -= 4 * mm
        c.line(15 * mm, y, 200 * mm, y)
        y -= 6 * mm

        c.setFont("Helvetica", 9)
        for r in receitas_por_grupo:
            ensure_space()
            c.drawString(15 * mm, y, (r["group__name"] or "")[:45])
            c.drawRightString(200 * mm, y, f"{(r['total'] or 0):.2f}")
            y -= 5 * mm

        y -= 6 * mm
        ensure_space()

        c.setFont("Helvetica-Bold", 11)
        c.drawString(15 * mm, y, "Despesas por Grupo")
        y -= 7 * mm
        c.setFont("Helvetica-Bold", 9)
        c.drawString(15 * mm, y, "Grupo")
        c.drawRightString(200 * mm, y, "Total (R$)")
        y -= 4 * mm
        c.line(15 * mm, y, 200 * mm, y)
        y -= 6 * mm

        c.setFont("Helvetica", 9)
        for r in despesas_por_grupo:
            ensure_space()
            c.drawString(15 * mm, y, (r["group__name"] or "")[:45])
            c.drawRightString(200 * mm, y, f"{(r['total'] or 0):.2f}")
            y -= 5 * mm

        c.save()
        buffer.seek(0)

        resp = HttpResponse(buffer.getvalue(), content_type="application/pdf")
        resp["Content-Disposition"] = f'attachment; filename="dre_{d1}_a_{d2}.pdf"'
        return resp

    return render(request, "finance/report_dre.html", {
        "d1": d1, "d2": d2,
        "receita_total": receita_total,
        "despesa_total": despesa_total,
        "resultado": resultado,
        "receitas_por_grupo": receitas_por_grupo,
        "despesas_por_grupo": despesas_por_grupo,
    })